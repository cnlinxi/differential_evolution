import openpyxl
import string
import testFunctions
from openpyxl.utils import get_column_letter
from deBase import DERand1Bin, DECurrentToPBest1Bin
from jade import JADEWithArchive as JADE
from sade import SaDE
from jde import jDE
from DE_RL import rlde
from hybrids import LocalJADE, HybridJADE, HybridjDE, sadJADE
from mixins import LoggingMixin, ParallelCostMixin, ValueToReachMixin
import time
import os
import csv
import sys
import multiprocessing

"""
This file tests the various options of the DifferentialEvolution class by
running them repeatedly on the test functions defined in testFunctions folder.

The results are exported to Microsoft Excel.
"""

# ALPHABET = string.uppercase

'''python study.py <repeat_num> --problem <problem_id> --file --parallel
<repeat_num>:int, 训练次数
<problem_id>:f1-10d/f1-30d/f2-10d/..., 本文件中problems:dict 中的键
--file 将运行信息输出到study.out文件中
--parallel 使用CPU并行
'''

def study():
    if "--file" in sys.argv:
        f = open('study.out', 'w')
    # algorithms = [DERand1Bin, jDE, SaDE, JADE, rlde]
    algorithms = [rlde]
    # algorithms=[dqde]
    repeats = int(sys.argv[1]) # 第一个参数，必须参数
    # Initialise Excel workbook
    wb = openpyxl.Workbook()
    file_id=str(time.strftime('%dd_%mmon_%Y_%Hh_%Mm'))
    wb_name = f'DE_Tests_{file_id}.xlsx'
    print('wb_name:',wb_name)
    worksheets = {}
    # Run the tests
    problems = {
        'f1-10d': testFunctions.sphere10d,
        'f1-30d': testFunctions.sphere30d,
        'f2-10d': testFunctions.rosenbrock10d,
        'f2-30d': testFunctions.rosenbrock30d,
        'f3-10d': testFunctions.step10d,
        'f3-30d': testFunctions.step30d,
        'f4-10d': testFunctions.ackley10d,
        'f4-30d': testFunctions.ackley30d,
        'f5-5d': testFunctions.shekel5d,
    }
    if "--problem" in sys.argv:
        i = sys.argv.index("--problem")
        p = sys.argv[i+1]
        problems = {p: problems[p]}
    for problem_descr, problem in sorted(problems.items()):
        print('Testing %s'%(problem_descr))
        for Algorithm in algorithms:
            problem_id = '%s_%s'%(problem_descr, Algorithm.__name__)
            class DE(ValueToReachMixin, LoggingMixin, Algorithm):
                pass
            if '30d' in problem_id:
                mfe = 100000 # 在30d上运行1e5
                if "--parallel" in sys.argv:
                    class DE(ParallelCostMixin, ValueToReachMixin, LoggingMixin, Algorithm):
                        pass
                    print('\nParallel Processing on %s CPUs\n'%(multiprocessing.cpu_count()))
                    if "--file" in sys.argv:
                        f.write('\nParallel Processing on %s CPUs\n\n'%(multiprocessing.cpu_count()))
            else:
                mfe = 50000 # 在其它问题维数运行5e4
            worksheets[problem_id] = wb.create_sheet()
            ws = worksheets[problem_id]
            ws.title = problem_id
            failures = 0
            for i in range(repeats):
                # Run the optimisation
                uuid = str(problem_id) + str(i) # 由problem_id(=problem_descripion+algorithm)+i(第i次训练)
                de = DE(costFile=problem, uuid=uuid, valueToReach=1e-6, maxFunctionEvals=mfe)
                run_name = '- Run %s with %s'%(i+1, problem_id)
                print(run_name)
                bestVector = de.optimise()
                print('bestVector:',bestVector)
                if "--file" in sys.argv:
                    f.write('%s\n%s\n\n'%(run_name, bestVector))
                # Dump convergence history
                with open(uuid + '.csv', 'r') as csvfile:
                    reader = csv.reader(csvfile)
                    for row_index, row in enumerate(reader):
                        for column_index, cell in enumerate(row):
                            column_letter = get_column_letter((4 * i) + (column_index + 1))
                            ws.cell('%s%s'%(column_letter, (row_index + 1))).value = cell
                os.remove(uuid + '.csv')
                del de

    print('Writing results to Excel...')
    # Remove the default sheet
    ws = wb.get_sheet_by_name('Sheet')
    wb.remove_sheet(ws)
    wb.save(wb_name)

if __name__ == '__main__':
    study()
